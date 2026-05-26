"""
SmartClinics AI – UAT Report Generator
Generates SmartClinics_AI_UAT_Report.xlsx based on static code analysis
of all 7 pages: index.html, dashboard.html, patient_lobby.html,
receptionist.html, translator.html, analytics.html, settings.html
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = r"C:\Partitions1\Ateequr Projects Healthcare\SmartclinicAI\SmartClinics_AI_UAT_Report.xlsx"

# ── Colour palette ──────────────────────────────────────────────
HEADER_BG   = "1558B0"   # brand blue
HEADER_FG   = "FFFFFF"
PASS_BG     = "D6F4DC"   # soft green
PASS_FG     = "1A7035"
FAIL_BG     = "FDDEDE"   # soft red
FAIL_FG     = "B03228"
PARTIAL_BG  = "FFF3CD"   # soft amber
PARTIAL_FG  = "856404"
ROW_ALT     = "F0F4FF"   # alternating row tint
BORDER_CLR  = "CCCCCC"

thin = Side(style="thin", color=BORDER_CLR)
medium = Side(style="medium", color="AAAAAA")
thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
header_border = Border(left=medium, right=medium, top=medium, bottom=medium)

def hfill(hex_str): return PatternFill("solid", fgColor=hex_str)
def hfont(bold=False, colour="000000", size=10): return Font(bold=bold, color=colour, size=size, name="Calibri")
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Test case data ──────────────────────────────────────────────
# Columns: ID, Module, Description, Expected, Actual, Status, Remarks
test_cases = [

    # ════════════════  INDEX.HTML  ════════════════
    ("TC-001", "index.html",
     "Navbar – anchor links (Features / How It Works / Pricing / Portals / About) navigate to correct sections",
     "Clicking each link smoothly scrolls to the corresponding section on the page",
     "All five anchor links contain correct href values (#features, #problem, #pricing, #portals, #about); smooth scroll enabled via CSS 'html { scroll-behavior: smooth }'",
     "Pass", "All in-page anchor IDs are present in the HTML"),

    ("TC-002", "index.html",
     "Navbar – 'Launch App' CTA button navigates to dashboard",
     "User is taken to app/dashboard.html",
     "href='app/dashboard.html' is correctly set on the Launch App button",
     "Pass", ""),

    ("TC-003", "index.html",
     "Hero – 'Start Free 14-Day Trial' button links to dashboard",
     "Clicking redirects to app/dashboard.html",
     "href='app/dashboard.html' confirmed on CTA anchor",
     "Pass", ""),

    ("TC-004", "index.html",
     "Hero – 'Watch Demo Video' button opens translator in demo mode",
     "Opens translator.html with lang=Spanish and patient=Demo+Patient query params",
     "href='app/translator.html?lang=Spanish&patient=Demo+Patient' confirmed",
     "Pass", ""),

    ("TC-005", "index.html",
     "Navbar – mobile hamburger menu toggles nav links",
     "Menu icon button shows/hides nav links on small screens",
     "JavaScript event listener on #menuBtn toggles 'active' class on #navLinks; CSS media query hides nav below 768 px",
     "Pass", ""),

    ("TC-006", "index.html",
     "Compliance badges render correctly in trust section",
     "HIPAA / HITRUST / SOC 2 / GDPR / WCAG / EPA badges visible with icons",
     "All six compliance badges present as static HTML with Font Awesome icons",
     "Pass", ""),

    ("TC-007", "index.html",
     "Hero video mockup translation bubbles animate",
     "Two translation bubbles fade in/out on the split-screen mockup",
     "CSS @keyframes fadeInOut applied to .translation-bubble elements; offset delays create alternating animation",
     "Pass", ""),

    ("TC-008", "index.html",
     "Responsive layout – single-column below 992 px, two-column above",
     "Hero grid stacks vertically on tablet/mobile; stats centre-align",
     "CSS media query at 992 px sets grid-template-columns:1fr; 768 px hides nav, stacks steps column",
     "Pass", ""),

    ("TC-009", "index.html",
     "Accessibility – skip-to-main-content link visible on keyboard focus",
     "Pressing Tab from browser URL bar surfaces a 'Skip to main content' link",
     ".skip-link moves from top:-9999px to top:0 on :focus; correct href='#main-content' linking to <main id='main-content'>",
     "Pass", ""),

    # ════════════════  DASHBOARD.HTML  ════════════════
    ("TC-010", "dashboard.html",
     "Sidebar navigation – all seven links render and point to correct pages",
     "Dashboard / Patients / Appointments / Translation History / Analytics / Settings / Receptionist View all navigate correctly",
     "All hrefs confirmed: dashboard.html, patients.html, appointments.html, history.html, analytics.html, settings.html, receptionist.html",
     "Pass", "Active state correctly set on Dashboard link"),

    ("TC-011", "dashboard.html",
     "Sidebar logo link returns to dashboard",
     "Clicking the logo navigates to dashboard.html",
     "href='dashboard.html' on .sidebar-logo anchor",
     "Pass", ""),

    ("TC-012", "dashboard.html",
     "'New Translation Session' button opens modal",
     "Modal overlay appears with slide-up animation; fields are pre-cleared",
     "openSessionModal() sets modal display:flex, opacity:1, adds .active class to content with 10ms delay",
     "Pass", ""),

    ("TC-013", "dashboard.html",
     "Session modal – close (×) button dismisses modal",
     "Clicking × fades out and hides modal",
     "closeModal() sets opacity:0, removes .active, then sets display:none after 300ms transition",
     "Pass", ""),

    ("TC-014", "dashboard.html",
     "Session modal – Cancel button dismisses modal",
     "Clicking Cancel closes modal without navigating",
     "#cancelBtn calls closeModal() via event listener",
     "Pass", ""),

    ("TC-015", "dashboard.html",
     "Session modal – clicking outside overlay closes modal",
     "Clicking the dark backdrop closes modal",
     "modal.addEventListener('click') checks e.target === modal before calling closeModal()",
     "Pass", ""),

    ("TC-016", "dashboard.html",
     "Session modal – 'Create Link & Start' submits and navigates to translator",
     "Redirects to translator.html with correct lang and patient query parameters",
     "initiateSession() reads #patientLang and #sessionPatientName; sets window.location.href to translator.html?lang=...&patient=...",
     "Pass", ""),

    ("TC-017", "dashboard.html",
     "Quick-action banner 'Prepare Session' button pre-fills modal",
     "Modal opens with Spanish selected and 'Maria Gonzalez (9:30 AM)' in patient name field",
     "onclick calls openSessionModal('Spanish', 'Maria Gonzalez (9:30 AM)'); both fields pre-populated",
     "Pass", ""),

    ("TC-018", "dashboard.html",
     "Schedule table 'Start Translation' buttons open modal pre-filled per patient",
     "Each button opens modal with the correct patient name and Spanish language",
     "onclick='openSessionModal(\"Spanish\", \"Maria Gonzalez\")' confirmed; button renders correctly",
     "Pass", ""),

    ("TC-019", "dashboard.html",
     "EHR Export buttons in schedule table trigger export action",
     "Clicking EHR button should initiate export process or show confirmation",
     "Buttons render with class 'btn btn-outline' and icon but have NO onclick handler or href defined",
     "Fail", "Bug: EHR export buttons are non-functional; no click handler attached"),

    ("TC-020", "dashboard.html",
     "'View Full Schedule' link navigates to full schedule",
     "Clicking link opens the full schedule view",
     "href='#' – placeholder only; no target page or navigation defined",
     "Fail", "Bug: Dead link – href='#' causes page jump to top with no navigation"),

    ("TC-021", "dashboard.html",
     "Header search bar accepts input",
     "User can type in the search field",
     "Input element renders correctly; no JavaScript search logic is implemented (prototype only)",
     "Partial", "UI only – search has no backend or filtering logic"),

    ("TC-022", "dashboard.html",
     "Language Settings icon button is functional",
     "Clicking globe/language icon opens language settings or overlay",
     "Button renders with title='Language Settings' but has no onclick handler",
     "Fail", "Bug: No action on Language Settings icon button"),

    ("TC-023", "dashboard.html",
     "Notifications bell button shows badge and is actionable",
     "Badge shows count (3); clicking opens notifications panel",
     "Bell button renders with badge '3' but has no onclick handler",
     "Fail", "Bug: No action on notification bell – badge is decorative only"),

    ("TC-024", "dashboard.html",
     "Stats cards display correct data",
     "12 Translations, 4.5h Saved, 325ms Latency, 8 EHR Exports Pending all visible",
     "All four stat-card elements render with correct values in HTML",
     "Pass", ""),

    ("TC-025", "dashboard.html",
     "Patient schedule table data renders correctly",
     "Four patients (Maria Gonzalez, Wei Chen, Amir Al-Fayed, Elena Rostova) with correct details",
     "Table rows statically defined with correct DOB, ID, time, language, and status",
     "Pass", ""),

    # ════════════════  PATIENT_LOBBY.HTML  ════════════════
    ("TC-026", "patient_lobby.html",
     "Page loads with correct session header",
     "Header shows 'Dr. Sarah Chen' and 'Connected (Secure)' pulsing status badge",
     "Static HTML header renders; .pulse-dot animation applies correctly via CSS @keyframes pulse",
     "Pass", ""),

    ("TC-027", "patient_lobby.html",
     "Doctor greeting message auto-appears after page load",
     "After ~1.5 s, a Spanish greeting bubble appears in the transcript area",
     "setTimeout( appendDocMsg('Hola…'), 1500 ) confirmed in JavaScript",
     "Pass", ""),

    ("TC-028", "patient_lobby.html",
     "Mic button – press and hold activates recording state",
     "Button turns red, ripple animation plays, helper text reads 'Listening…'",
     "mousedown/touchstart sets recording=true, adds .active class, updates helperText",
     "Pass", ""),

    ("TC-029", "patient_lobby.html",
     "Mic button – release simulates translation and patient response",
     "Helper text shows 'Translating…', then patient bubble appears; doctor follow-up after 3 s",
     "stopRecord() sets state, calls appendPatientMsg after 800ms, then appendDocMsg after 3000ms",
     "Pass", ""),

    ("TC-030", "patient_lobby.html",
     "URL param 'lang' is read and applied to patient speaker label",
     "If lang=Mandarin, patient bubble displays 'You (Mandarin)'",
     "const myLang = qs.get('lang') || 'Spanish'; used in appendPatientMsg speaker label",
     "Pass", ""),

    ("TC-031", "patient_lobby.html",
     "End Session button (phone-slash) closes the window",
     "Clicking phone-slash icon closes the current browser tab/window",
     "onclick='window.close()' on end button",
     "Pass", "Expected behaviour for a patient popup/tab"),

    ("TC-032", "patient_lobby.html",
     "Transcript area scrolls to latest message automatically",
     "New messages scroll into view without manual scroll",
     "scrollToBottom() sets transcript.scrollTop = transcript.scrollHeight after each append",
     "Pass", ""),

    ("TC-033", "patient_lobby.html",
     "Responsive / mobile-first layout",
     "Page fills full viewport height; controls fixed at bottom; transcript scrollable in middle",
     "viewport meta has maximum-scale=1.0 user-scalable=no; body uses flex-direction:column with overflow:hidden",
     "Pass", "Designed specifically as mobile-first patient view"),

    # ════════════════  RECEPTIONIST.HTML  ════════════════
    ("TC-034", "receptionist.html",
     "Back to Provider Portal link navigates to dashboard",
     "Clicking '← Provider Portal' returns to dashboard.html",
     "href='dashboard.html' on back link confirmed",
     "Pass", ""),

    ("TC-035", "receptionist.html",
     "Live clock displays and updates every second",
     "Current time shown in HH:MM:SS format, updating in real time",
     "updateClock() called once then setInterval(updateClock, 1000); uses toLocaleTimeString",
     "Pass", ""),

    ("TC-036", "receptionist.html",
     "Queue list renders all 6 patients on load",
     "Queue shows Elena Rostova, Amir Al-Fayed, Maria Gonzalez, Wei Chen, Priya Sharma, Carlos Mendez",
     "renderQueue() called at bottom of script; maps queueData array to HTML rows",
     "Pass", ""),

    ("TC-037", "receptionist.html",
     "Queue count badge shows correct number",
     "Badge displays count of non-'done' patients",
     "queueData.filter(q => q.status !== 'done').length used to update #queueCount",
     "Pass", ""),

    ("TC-038", "receptionist.html",
     "Clicking a queue item selects patient and populates workspace",
     "Patient name, meta, appointment, insurance, phone, room, copay all load in right panel",
     "selectPatient(id) hides #noSelection, shows #patientWorkspace, sets all fields from queueData",
     "Pass", ""),

    ("TC-039", "receptionist.html",
     "Status badge in workspace reflects patient check-in status correctly",
     "'✓ Checked In' for status:here, '📍 In Room' for called, '⏳ Scheduled' for waiting",
     "statusMap and statusText objects correctly mapped; badge class updated",
     "Pass", ""),

    ("TC-040", "receptionist.html",
     "Wait time calculated and displayed from check-in time",
     "Elapsed minutes since check-in shown; '—' for patients not yet checked in",
     "Time parsed from checkinTime string; Math.floor((now - checkin)/60000) computed",
     "Pass", ""),

    ("TC-041", "receptionist.html",
     "'Start Translation Session' button navigates to translator with correct params",
     "Opens translator.html?lang=<patientLang>&patient=<patientName>",
     "startSelectedSession() constructs URL from selected patient's lang and name",
     "Pass", ""),

    ("TC-042", "receptionist.html",
     "'Call to Exam Room' button assigns room and updates patient status",
     "Prompts for room name; updates patient status to 'called' and shows room in details",
     "callPatient() uses window.prompt(); updates p.room and p.status, refreshes view",
     "Pass", "Uses browser prompt – acceptable for prototype"),

    ("TC-043", "receptionist.html",
     "Assign Room quick action updates patient record",
     "Prompt input assigns room and sets status to 'called'",
     "assignRoom() with prompt; updates queueData and re-selects patient",
     "Pass", ""),

    ("TC-044", "receptionist.html",
     "Collect Copay action shows confirmation",
     "Alert displays copay amount and patient name",
     "collectCopay() calls alert() with patient copay and name",
     "Pass", "Simulated – no payment gateway in prototype"),

    ("TC-045", "receptionist.html",
     "Send Patient Link action sends SMS confirmation",
     "Confirm dialog shows phone + link; on confirm, alert shows success",
     "sendPatientLink() uses confirm() and alert(); constructs patient_lobby.html URL",
     "Pass", ""),

    ("TC-046", "receptionist.html",
     "View Records quick action navigates to patients page",
     "Clicking redirects to patients.html",
     "viewRecords() sets window.location.href = 'patients.html'",
     "Pass", ""),

    ("TC-047", "receptionist.html",
     "Book Follow-Up quick action navigates to appointments page",
     "Clicking redirects to appointments.html",
     "bookFollowup() sets window.location.href = 'appointments.html'",
     "Pass", ""),

    ("TC-048", "receptionist.html",
     "Mark No-Show removes patient from queue",
     "Confirm dialog; on confirm patient removed from view and queue count decreases",
     "markNoShow() sets p.status='done', hides workspace, calls renderQueue()",
     "Pass", ""),

    ("TC-049", "receptionist.html",
     "SMS template buttons populate textarea",
     "Clicking 'Room Ready', 'Running Late', 'Appt Reminder', 'Lab Results' fills SMS body",
     "All four sms-template-btn buttons call setSMS(text) which sets smsBody.value",
     "Pass", ""),

    ("TC-050", "receptionist.html",
     "Send SMS validates empty message",
     "Alert shown if textarea is blank when Send SMS clicked",
     "sendSMS() checks body.trim() and calls alert('Please enter a message.')",
     "Pass", ""),

    ("TC-051", "receptionist.html",
     "Send SMS with content shows success and clears textarea",
     "Alert confirms SMS sent to patient phone; body field is cleared",
     "sendSMS() alerts success, then sets smsBody.value = ''",
     "Pass", ""),

    ("TC-052", "receptionist.html",
     "Walk-In Check-In modal opens from header button",
     "Modal slides up with correct form fields",
     "openCheckinModal() sets display:flex, opacity:1, adds .active class",
     "Pass", ""),

    ("TC-053", "receptionist.html",
     "Walk-In Check-In – patient name required validation",
     "Alert shown if name field is empty on submission",
     "completeCheckin() checks if(!name) and alerts",
     "Pass", ""),

    ("TC-054", "receptionist.html",
     "Walk-In Check-In – complete adds patient to queue",
     "New entry appears in queue list with correct data",
     "completeCheckin() creates new entry object and pushes to queueData; renderQueue() called",
     "Pass", ""),

    ("TC-055", "receptionist.html",
     "Receptionist page responsive layout on mobile",
     "Queue panel and workspace stack vertically on small screens",
     "Fixed 380 px queue panel width with no CSS media queries; layout breaks on screens < 760 px",
     "Fail", "Bug: No responsive breakpoints defined for the receptionist dual-panel layout"),

    # ════════════════  TRANSLATOR.HTML  ════════════════
    ("TC-056", "translator.html",
     "Header displays patient name from URL parameter",
     "Page title shows patient name passed via ?patient= query string",
     "params.get('patient') || 'Maria Gonzalez' set as innerText of #headerPatientName",
     "Pass", ""),

    ("TC-057", "translator.html",
     "Language flag in transcript header updates from URL lang param",
     "Flag abbreviation matches language (ES for Spanish, ZH for Mandarin, etc.)",
     "langMap object maps language names to 2-char codes; #targetLangFlag updated",
     "Pass", ""),

    ("TC-058", "translator.html",
     "Session timer counts up from 00:00",
     "Timer in header increments every second throughout the session",
     "setInterval increments seconds, pads to MM:SS, updates #sessionTimer",
     "Pass", ""),

    ("TC-059", "translator.html",
     "Ping counter fluctuates dynamically",
     "Latency value changes every 3 seconds between ~180–230 ms; colour changes based on threshold",
     "setInterval every 3000ms; random value 180-230; metric-good vs metric-warn class applied",
     "Pass", ""),

    ("TC-060", "translator.html",
     "Doctor greeting message appears after page load",
     "Chat bubble with Dr. Chen's English + translated greeting renders after ~1.5 s",
     "setTimeout( addMessageToChat('doc', ...), 1500 ) in DOMContentLoaded",
     "Pass", ""),

    ("TC-061", "translator.html",
     "Patient response appears after delay",
     "Simulated patient reply in target language appears ~5.5 s after page load",
     "setTimeout( addMessageToChat('patient', ...), 5500 ) in DOMContentLoaded",
     "Pass", ""),

    ("TC-062", "translator.html",
     "Emergency banner hidden by default; shown on keyword detection",
     "Banner initially hidden; appears when emergency keywords ('chest pain') detected",
     "#emergencyBanner has display:none in CSS; checkForEmergency() sets display:flex when triggered by translatedText",
     "Pass", "Emergency detection fires on patient response containing 'chest pain'"),

    ("TC-063", "translator.html",
     "Demo chat form – typing and submitting adds doctor message to transcript",
     "Typed text appears as a doctor bubble with mock translation below",
     "handleDemoSubmit() in app.js reads #demoInput, calls addMessageToChat('doc', …)",
     "Pass", ""),

    ("TC-064", "translator.html",
     "Video call control buttons (microphone, camera) toggle active state",
     "Clicking mic or camera buttons toggles their active visual state",
     "Buttons render with class 'control-btn active' but have NO onclick handlers defined; state does not change on click",
     "Fail", "Bug: Mic and camera buttons are visually present but non-interactive"),

    ("TC-065", "translator.html",
     "Share Screen and Interpreter Settings control buttons are interactive",
     "Clicking opens screen share or interpreter settings",
     "Buttons render but have NO onclick handlers defined",
     "Fail", "Bug: Share Screen and Interpreter Settings buttons have no actions"),

    ("TC-066", "translator.html",
     "'End Session' button navigates back to dashboard",
     "Clicking End Session returns user to dashboard.html",
     "onclick='window.location.href=\"dashboard.html\"' confirmed on end-call button",
     "Pass", ""),

    ("TC-067", "translator.html",
     "Back arrow in header navigates to dashboard",
     "Clicking ← icon returns to dashboard.html",
     "href='dashboard.html' on back arrow anchor",
     "Pass", ""),

    ("TC-068", "translator.html",
     "'Open Patient View' button in notification banner opens patient lobby",
     "Patient lobby opens in new tab with correct lang param",
     "window.open() called with patient_lobby.html?lang=<lang> in a new tab",
     "Pass", ""),

    ("TC-069", "translator.html",
     "Close button on patient link notification hides the banner",
     "Clicking × dismisses the notification bar",
     "onclick sets #patientLinkNotification style.display='none'",
     "Pass", ""),

    ("TC-070", "translator.html",
     "Live transcript panel auto-scrolls to latest message",
     "New messages scroll into view without manual scrolling",
     "container.scrollTop = container.scrollHeight in addMessageToChat()",
     "Pass", ""),

    # ════════════════  ANALYTICS.HTML  ════════════════
    ("TC-071", "analytics.html",
     "Sidebar navigation links work correctly",
     "All nav links navigate to correct pages; Analytics marked as active",
     "All hrefs confirmed; analytics.html nav-item has class='active'",
     "Pass", ""),

    ("TC-072", "analytics.html",
     "KPI cards display correct metric values",
     "342 translations, 98.2% accuracy, 143h saved, $4,290 savings all visible",
     "All four .kpi-card elements with correct static values present in HTML",
     "Pass", ""),

    ("TC-073", "analytics.html",
     "Translation Volume line chart renders",
     "Line chart showing 30-day translation volume appears",
     "Chart.js line chart instantiated on #volumeChart canvas with 30 data points",
     "Pass", ""),

    ("TC-074", "analytics.html",
     "Language Distribution doughnut chart renders",
     "Doughnut chart showing language split (Spanish 48%, Mandarin 22%, etc.) renders",
     "Chart.js doughnut chart on #langChart with 6 language segments",
     "Pass", ""),

    ("TC-075", "analytics.html",
     "AI Model Usage bar chart renders",
     "Bar chart showing OpenAI / Gemini / Bedrock usage renders",
     "Chart.js bar chart on #modelChart with 3 bars",
     "Pass", ""),

    ("TC-076", "analytics.html",
     "Language breakdown progress bars display correctly",
     "Five horizontal bars (Spanish 48% → Russian 6%) with correct widths and colours",
     "Static CSS width percentages on .lang-bar-fill elements match labels",
     "Pass", ""),

    ("TC-077", "analytics.html",
     "Patient satisfaction rating block displays 4.8/5 stars",
     "Star rating, score, breakdown (72%/20%/8%), and response count visible",
     "Static HTML with fas fa-star icons and progress bars rendered correctly",
     "Pass", ""),

    ("TC-078", "analytics.html",
     "ROI Summary card displays four financial metrics",
     "$6,150 interpreter cost, $199 AI cost, $5,951 net savings, 96.8% reduction shown",
     "Static .roi-stat elements inside .roi-card present with correct values",
     "Pass", ""),

    ("TC-079", "analytics.html",
     "Date filter buttons toggle active state",
     "Clicking '7 Days', '30 Days', or 'Year' highlights that button; others de-highlight",
     "setRange() removes 'active' from all buttons and adds to clicked button",
     "Pass", "Note: chart data does NOT re-render on filter change – prototype limitation"),

    ("TC-080", "analytics.html",
     "'Download Report' button triggers download simulation",
     "Alert confirms PDF download with filename analytics_report_may2026.pdf",
     "downloadReport() calls alert() with download message",
     "Pass", "Simulated download – no actual file generated"),

    # ════════════════  SETTINGS.HTML  ════════════════
    ("TC-081", "settings.html",
     "Sidebar navigation links work correctly",
     "All nav links navigate to correct pages; Settings marked as active",
     "All hrefs confirmed; settings.html nav-item has class='active'",
     "Pass", ""),

    ("TC-082", "settings.html",
     "Settings tab navigation switches panels correctly",
     "Clicking Profile/Translation/Notifications/Integrations/Security/Billing/Team shows correct panel",
     "switchTab() removes/adds 'active' class on .settings-nav-item and .settings-panel elements",
     "Pass", ""),

    ("TC-083", "settings.html",
     "'Save Changes' button shows success feedback",
     "Button turns green with checkmark for 2 seconds then resets",
     "saveSettings() updates button innerHTML and background; setTimeout resets after 2000ms",
     "Pass", ""),

    ("TC-084", "settings.html",
     "Profile tab – form fields are editable",
     "First/Last Name, Title, NPI, Email, Phone, Bio fields accept user input",
     "All inputs pre-filled with values; textarea editable; no readonly attributes",
     "Pass", ""),

    ("TC-085", "settings.html",
     "Profile tab – 'Upload New Photo' button triggers file upload",
     "Clicking prompts file chooser for new avatar image",
     "Button renders with class 'btn btn-outline' but has NO onclick handler or file input connected",
     "Fail", "Bug: Upload Photo button is non-functional – no file input or handler"),

    ("TC-086", "settings.html",
     "Translation tab – toggle switches function correctly",
     "Live Subtitles, Speaker Diarization, Auto Transcription, Emergency SMS toggles can be toggled",
     "CSS .toggle-switch with hidden checkbox and .slider work via native browser checkbox toggle",
     "Pass", ""),

    ("TC-087", "settings.html",
     "Translation tab – 'Add Term' adds a new glossary row",
     "Clicking 'Add Term' appends a new empty row with two inputs and a delete button",
     "addGlossaryRow() creates div with two form-control inputs and delete button; appended to #glossaryList",
     "Pass", ""),

    ("TC-088", "settings.html",
     "Translation tab – delete (trash) button removes glossary row",
     "Clicking trash icon removes that row from the glossary list",
     "onclick='this.parentElement.remove()' on delete button; removes parent div",
     "Pass", ""),

    ("TC-089", "settings.html",
     "Notifications tab – all six toggle switches function",
     "Each notification toggle can be enabled/disabled independently",
     "Six .toggle-row elements with toggle-switch components; all use native checkbox mechanism",
     "Pass", ""),

    ("TC-090", "settings.html",
     "Integrations tab – Epic 'Configure' button performs an action",
     "Clicking Configure opens Epic EHR configuration panel or modal",
     "Button renders with class 'btn btn-outline' but has NO onclick handler",
     "Fail", "Bug: Epic Configure button has no action"),

    ("TC-091", "settings.html",
     "Integrations tab – Cerner and athenahealth 'Connect' buttons initiate connection",
     "Clicking Connect opens OAuth flow or configuration modal",
     "Both Connect buttons render with class 'btn btn-primary' but have NO onclick handlers",
     "Fail", "Bug: EHR Connect buttons (Cerner, athenahealth) are non-functional"),

    ("TC-092", "settings.html",
     "Security tab – password fields accept input",
     "Current, New, and Confirm password fields render and accept input",
     "Three <input type='password'> fields render without readonly attribute",
     "Pass", ""),

    ("TC-093", "settings.html",
     "Security tab – 2FA and security toggles function",
     "2FA, Biometric Login, Auto-Logout toggles can be switched",
     "Three toggle-row elements with toggle switches work via native checkbox",
     "Pass", ""),

    ("TC-094", "settings.html",
     "Security tab – 'Delete Account' button shows contact information",
     "Clicking shows alert directing user to contact support",
     "onclick='alert(\"This action would delete all data. Contact support@smartclinics.ai\")' confirmed",
     "Pass", ""),

    ("TC-095", "settings.html",
     "Billing tab – current plan details display correctly",
     "Professional plan at $199/mo with next billing date June 1, 2026 shown",
     "Static HTML .kpi values present with correct plan data",
     "Pass", ""),

    ("TC-096", "settings.html",
     "Billing tab – usage progress bars display correctly",
     "Translation Minutes (71%) and Active Sessions (68%) bars render",
     "Static CSS width on progress bar divs; 1423/2000 min and 342 sessions shown",
     "Pass", ""),

    ("TC-097", "settings.html",
     "Billing tab – 'Change Plan' button is functional",
     "Clicking opens plan upgrade/downgrade options",
     "Button renders with 'btn btn-outline' but has NO onclick handler",
     "Fail", "Bug: Change Plan button has no action"),

    ("TC-098", "settings.html",
     "Billing tab – 'Cancel Subscription' button is functional",
     "Clicking opens cancellation confirmation flow",
     "Button renders with danger styling but has NO onclick handler",
     "Fail", "Bug: Cancel Subscription button has no action"),

    ("TC-099", "settings.html",
     "Team tab – team member list renders with correct data",
     "Four members (Dr. Sarah Chen, Dr. James Washington, Maria Lopez, Alex Kim) displayed with roles/status",
     "HTML contains raw template literal strings '${buildTeamMember(...)}' in static HTML body – browser renders as literal text, NOT as actual member cards",
     "Fail", "Critical Bug: Template literals used in HTML markup instead of JavaScript; team members not rendered"),

    ("TC-100", "settings.html",
     "Team tab – 'Invite' button sends invitation",
     "Clicking Invite prompts for email and confirms invitation sent",
     "inviteTeam() calls window.prompt() for email; if entered, calls alert() with confirmation",
     "Pass", ""),
]

def create_report():
    wb = openpyxl.Workbook()

    # ── Cover sheet ─────────────────────────────────────────────
    cover = wb.active
    cover.title = "Cover"
    cover.column_dimensions["A"].width = 60

    cover_data = [
        ("SmartClinics AI – User Acceptance Testing Report", True, 18),
        ("",                                                  False, 11),
        ("Project Name:  SmartClinics AI",                  False, 11),
        ("Version:       1.0 (Prototype)",                   False, 11),
        ("Test Date:     2026-05-09",                         False, 11),
        ("Tester:        Claude (Automated Code Analysis)",   False, 11),
        ("Scope:         index.html | dashboard.html | patient_lobby.html |", False, 11),
        ("               receptionist.html | translator.html | analytics.html | settings.html", False, 11),
        ("",                                                  False, 11),
        ("SUMMARY",                                           True, 13),
        ("",                                                  False, 11),
    ]

    total    = len(test_cases)
    passes   = sum(1 for t in test_cases if t[5] == "Pass")
    fails    = sum(1 for t in test_cases if t[5] == "Fail")
    partials = sum(1 for t in test_cases if t[5] == "Partial")

    cover_data += [
        (f"Total Test Cases:  {total}",             False, 11),
        (f"Passed:            {passes}  ({passes*100//total}%)",  False, 11),
        (f"Failed:            {fails}  ({fails*100//total}%)",    False, 11),
        (f"Partial:           {partials}  ({partials*100//total}%)", False, 11),
    ]

    for row_data in cover_data:
        text, bold, size = row_data
        cell = cover.cell(row=cover.max_row + 1 if cover.max_row > 0 else 1, column=1, value=text)
        cell.font = Font(bold=bold, size=size, name="Calibri",
                         color=HEADER_BG if bold else "333333")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # colour summary rows
    for row in range(12, 16):
        c = cover.cell(row=row, column=1)
        if "Passed" in (c.value or ""):
            c.font = Font(bold=True, size=11, color=PASS_FG, name="Calibri")
        elif "Failed" in (c.value or ""):
            c.font = Font(bold=True, size=11, color=FAIL_FG, name="Calibri")
        elif "Partial" in (c.value or ""):
            c.font = Font(bold=True, size=11, color=PARTIAL_FG, name="Calibri")

    # ── UAT Detail sheet ────────────────────────────────────────
    ws = wb.create_sheet("UAT Test Cases")
    headers = ["Test Case ID", "Module / Page", "Test Description",
               "Expected Result", "Actual Result", "Status", "Remarks"]
    col_widths = [13, 22, 40, 38, 42, 10, 38]

    # Header row
    for col, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=1, column=col, value=hdr)
        c.fill    = hfill(HEADER_BG)
        c.font    = hfont(bold=True, colour=HEADER_FG, size=11)
        c.alignment = center()
        c.border  = header_border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, tc in enumerate(test_cases, start=2):
        alt = (row_idx % 2 == 0)
        status = tc[5]

        if status == "Pass":
            status_fill = hfill(PASS_BG)
            status_font = hfont(bold=True, colour=PASS_FG)
        elif status == "Fail":
            status_fill = hfill(FAIL_BG)
            status_font = hfont(bold=True, colour=FAIL_FG)
        else:
            status_fill = hfill(PARTIAL_BG)
            status_font = hfont(bold=True, colour=PARTIAL_FG)

        row_fill = hfill(ROW_ALT) if alt else hfill("FFFFFF")

        for col, value in enumerate(tc, start=1):
            c = ws.cell(row=row_idx, column=col, value=value)
            c.border    = thin_border
            c.alignment = left() if col not in (1, 2, 6) else center()

            if col == 6:   # Status column
                c.fill   = status_fill
                c.font   = status_font
                c.alignment = center()
            else:
                c.fill = row_fill
                c.font = hfont(size=10)

        ws.row_dimensions[row_idx].height = 52

    # Auto-filter
    ws.auto_filter.ref = f"A1:G{len(test_cases)+1}"

    # ── Summary sheet ────────────────────────────────────────────
    ss = wb.create_sheet("Summary by Page")
    ss_headers = ["Module / Page", "Total Tests", "Passed", "Failed", "Partial", "Pass Rate"]
    ss_widths   = [25, 13, 13, 13, 13, 14]

    for col, (h, w) in enumerate(zip(ss_headers, ss_widths), start=1):
        c = ss.cell(row=1, column=col, value=h)
        c.fill = hfill(HEADER_BG); c.font = hfont(bold=True, colour=HEADER_FG, size=11)
        c.alignment = center(); c.border = header_border
        ss.column_dimensions[get_column_letter(col)].width = w
    ss.row_dimensions[1].height = 28

    pages = ["index.html", "dashboard.html", "patient_lobby.html",
             "receptionist.html", "translator.html", "analytics.html", "settings.html"]

    for r, page in enumerate(pages, start=2):
        pts = [t for t in test_cases if t[1] == page]
        tot = len(pts); p = sum(1 for t in pts if t[5]=="Pass")
        f = sum(1 for t in pts if t[5]=="Fail"); pa = sum(1 for t in pts if t[5]=="Partial")
        rate = f"{p*100//tot}%" if tot else "—"

        vals = [page, tot, p, f, pa, rate]
        for col, v in enumerate(vals, start=1):
            c = ss.cell(row=r, column=col, value=v)
            c.border = thin_border
            c.alignment = center()
            c.font = hfont(size=10)
            if col == 1:
                c.fill = hfill(ROW_ALT); c.alignment = left()
            elif col == 3:
                c.fill = hfill(PASS_BG); c.font = hfont(bold=True, colour=PASS_FG)
            elif col == 4:
                c.fill = hfill(FAIL_BG); c.font = hfont(bold=True, colour=FAIL_FG)
            elif col == 5:
                c.fill = hfill(PARTIAL_BG); c.font = hfont(bold=True, colour=PARTIAL_FG)
            elif col == 6:
                pct = int(rate.replace("%","")) if rate != "—" else 0
                if pct == 100: c.fill = hfill(PASS_BG); c.font = hfont(bold=True, colour=PASS_FG)
                elif pct >= 70: c.fill = hfill(PARTIAL_BG); c.font = hfont(bold=True, colour=PARTIAL_FG)
                else: c.fill = hfill(FAIL_BG); c.font = hfont(bold=True, colour=FAIL_FG)
            else:
                c.fill = hfill("FFFFFF") if r % 2 == 0 else hfill("FFFFFF")

    # Totals row
    tr = len(pages) + 2
    ss.cell(row=tr, column=1, value="TOTAL").font = hfont(bold=True, size=11)
    ss.cell(row=tr, column=1).fill = hfill("E8EFF8")
    ss.cell(row=tr, column=1).border = header_border
    ss.cell(row=tr, column=2, value=total).font = hfont(bold=True, size=11)
    ss.cell(row=tr, column=3, value=passes).font = hfont(bold=True, colour=PASS_FG, size=11)
    ss.cell(row=tr, column=3).fill = hfill(PASS_BG)
    ss.cell(row=tr, column=4, value=fails).font = hfont(bold=True, colour=FAIL_FG, size=11)
    ss.cell(row=tr, column=4).fill = hfill(FAIL_BG)
    ss.cell(row=tr, column=5, value=partials).font = hfont(bold=True, colour=PARTIAL_FG, size=11)
    ss.cell(row=tr, column=5).fill = hfill(PARTIAL_BG)
    ss.cell(row=tr, column=6, value=f"{passes*100//total}%").font = hfont(bold=True, size=11)
    for col in range(1, 7):
        ss.cell(row=tr, column=col).border = header_border
        ss.cell(row=tr, column=col).alignment = center()

    # ── Set sheet order ─────────────────────────────────────────
    wb.active = wb["UAT Test Cases"]

    wb.save(OUTPUT_PATH)
    print(f"Report saved: {OUTPUT_PATH}")
    print(f"Total: {total}  |  Pass: {passes}  |  Fail: {fails}  |  Partial: {partials}")

if __name__ == "__main__":
    create_report()
